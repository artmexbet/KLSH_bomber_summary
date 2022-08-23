import openpyxl as xl

filename = input("Введите путь до файла таблицы: ")
table_index = int(input("Введите номер листа в таблице (начиная с нуля): "))

wb = xl.load_workbook(filename)
sheet = wb[wb.sheetnames[table_index]]

bombers = {}

# Начинаем с двойки, потому что строки считаются с единицы, и мы не учитываем шапку
for i in range(2, sheet.max_row + 1):
    team = sheet.cell(i, 2).value  # Команда
    name = sheet.cell(i, 3).value  # Имя человека
    points = sheet.cell(i, 4).value
    if name in bombers.keys():
        bombers[name][1] += points
    else:
        bombers[name] = [team, points]

# Сортируем сначала по возрастанию баллов, потом по фамилии, потом по команде
bombers = sorted(bombers.items(), key=lambda x: (-x[1][1], x[0], x[1][0]))

filename = input("Введите путь (с названием и расширением (.xlsx) файла), куда нужно сохранить итоговый файл: ")
new_wb = xl.Workbook()
new_sheet = new_wb.create_sheet("cool_bombers")

new_sheet["A1"] = "Имя"
new_sheet["B1"] = "Команда"
new_sheet["C1"] = "Сумма"

count = 2

for name, others in bombers:
    new_sheet.cell(count, 1, name)
    new_sheet.cell(count, 2, others[0])
    new_sheet.cell(count, 3, others[1])
    count += 1

new_wb.save(filename)
print(f"""Нужные данные лежат в листе под названием 'cool_bombers'
Файл лежит в {filename}""")
input()
